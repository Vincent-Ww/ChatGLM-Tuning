# -- coding: utf-8 --
# @Author : wuzixun
# @Time : 2023/5/19 11:26 AM
# @File : llm_ft_evaluete_e2e.py

from time import time
import torch
from transformers import AutoTokenizer, AutoModel
from peft import PeftModel
from openpyxl import Workbook
import json
from tqdm import tqdm
import pandas as pd
import traceback
from q_matching import Retrieval

PEFT_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/output-question-20230520/"
CHATGLM_PATH = "/home/xiezizhe/wuzixun/LLM/chatglm-6b"
DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_h2h_question/resample_h2h_question_dev.json"


def chatglm_inference(model, tokenizer, sample):
    context = f"Instruction: {sample['instruction']}\n"
    context += f"Input: {sample['input']}\n"
    context += "Answer: "

    response, _ = model.chat(tokenizer, context, history=[])
    return response


if __name__ == "__main__":

    tokenizer = AutoTokenizer.from_pretrained(CHATGLM_PATH, trust_remote_code=True)
    model = AutoModel.from_pretrained(CHATGLM_PATH, load_in_8bit=True, trust_remote_code=True, device_map='auto')
    model = model.eval()

    model = PeftModel.from_pretrained(model, PEFT_PATH)

    with open(DEV_DATA_PATH, "r", encoding='utf-8') as f:
        dev_data = json.load(f)

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = '对话'
    sheet.cell(1, 2).value = '标准问'
    sheet.cell(1, 3).value = 'LLM结果(微调后)'
    sheet.cell(1, 4).value = 'FT'
    sheet.cell(1, 5).value = "LLM结果匹配人工Q"
    sheet.cell(1, 6).value = "LLM结果对应FT"
    sheet.cell(1, 7).value = "一级FT是否正确"
    sheet.cell(1, 8).value = "二级FT是否正确"

    with open("data/h2h_question/manual_q2ft.json", "r") as f:
        faq2ft = json.load(f)

    ret = Retrieval(load_from_disk=True, persist_directory=".chroma/biaozhunwen")

    nrow = 2
    for sample in tqdm(dev_data):
        try:
            sheet.cell(nrow, 1).value = sample['input']
            llm_answer = chatglm_inference(model, tokenizer, sample)
            sheet.cell(nrow, 2).value = sample['output']
            sheet.cell(nrow, 3).value = llm_answer

            multi_ft_label = sample['ft_name'].split("-")
            if len(multi_ft_label) >= 1:
                ft_1_label = multi_ft_label[0]
            if len(multi_ft_label) >= 2:
                ft_2_label = multi_ft_label[1]
            ft_1_2_label = ft_1_label + "-" + ft_2_label
            sheet.cell(nrow, 4).value = ft_1_2_label

            retrieved_q = ret.retrieve(llm_answer)
            sheet.cell(nrow, 5).value = retrieved_q

            ft = faq2ft[retrieved_q]
            sheet.cell(nrow, 6).value = ft


            def is_first_equal(ft_label, ft_predict):
                label = ft_label.split("-")[0]
                predict = ft_predict.split("-")[0]
                return label == predict


            def is_second_equal(ft_label, ft_predict):
                label = ft_label.split("-")
                predict = ft_predict.split("-")
                return label == predict


            sheet.cell(nrow, 7).value = is_first_equal(ft_1_2_label, ft)
            sheet.cell(nrow, 8).value = is_second_equal(ft_1_2_label, ft)

            nrow += 1
            if nrow % 200 == 0 and nrow != 0:
                workbook.save(f"人人对话LLM验证{nrow}_匹配人工Q.xlsx")
                print(f"人人对话LLM验证{nrow}_匹配人工Q.xlsx: Save!")
        except Exception as e:
            traceback.print_exc()
            print("Exception: ", e)
    workbook.save(f"人人对话LLM验证{nrow}_匹配人工Q.xlsx")
