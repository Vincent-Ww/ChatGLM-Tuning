# -- coding: utf-8 --
# @Author : wuzixun
# @Time : 2023/5/19 11:26 AM
# @File : llm_ft_evaluete_e2e.py

from transformers import AutoTokenizer, AutoModel
from peft import PeftModel
from openpyxl import Workbook
import json
from tqdm import tqdm
import traceback

PEFT_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/ks-ai-ft5-20230626"
CHATGLM_PATH = "/home/xiezizhe/wuzixun/LLM/chatglm-6b"
# DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_ai_ft3/ks_ai_nomanual_ft_format_12-25_dev.json"
# DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/tmp/ai_ft_format_filter0.json"
DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_ai_ft_anno/ks_ai_ft_0519-0601_anno_dev.json"
OUTPUT_PATH = "智能0626LLM验证_onAI{}.xlsx"


def chatglm_inference(model, tokenizer, sample):
    context = f"Instruction: {sample['instruction']}\n"
    context += f"Input: {sample['input']}\n"
    context += "Answer: "

    response, _ = model.chat(tokenizer, context, history=[])
    return response


if __name__ == "__main__":

    tokenizer = AutoTokenizer.from_pretrained(CHATGLM_PATH, trust_remote_code=True)
    model = AutoModel.from_pretrained(CHATGLM_PATH, load_in_8bit=True, trust_remote_code=True, device_map='auto')
    model = model.eval()

    model = PeftModel.from_pretrained(model, PEFT_PATH)

    with open(DEV_DATA_PATH, "r", encoding='utf-8') as f:
        dev_data = json.load(f)

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = "会话ID"
    sheet.cell(1, 2).value = '对话'
    sheet.cell(1, 3).value = '首次关联工单功能树'
    sheet.cell(1, 4).value = 'LLM结果(微调后)'
    sheet.cell(1, 5).value = "首次关联工单功能树和预测FT一致(一级FT)"
    sheet.cell(1, 6).value = "首次关联工单功能树和预测FT一致(二级FT)"
    sheet.cell(1, 7).value = "mannual_reason"
    sheet.cell(1, 8).value = "是否转接"
    sheet.cell(1, 9).value = "转接后技能组"
    sheet.cell(1, 10).value = "线上预测一级FT"
    sheet.cell(1, 11).value = "线上预测二级FT"
    sheet.cell(1, 12).value = "AI一级FT是否正确"
    sheet.cell(1, 13).value = "AI二级FT是否正确"
    sheet.cell(1, 14).value = "route_source"
    sheet.cell(1, 15).value = "是否标注"
    sheet.cell(1, 16).value = "标注前结果"

    with open("data/manual_ft_q_map/manual_q2ft.json", "r") as f:
        faq2ft = json.load(f)

    nrow = 2
    for sample in tqdm(dev_data):
        try:
            sheet.cell(nrow, 1).value = sample['session_id']
            sheet.cell(nrow, 2).value = sample['input']
            llm_answer = chatglm_inference(model, tokenizer, sample)
            ft_label = sample['output']
            sheet.cell(nrow, 3).value = ft_label
            sheet.cell(nrow, 4).value = llm_answer


            def is_first_equal(ft_label, ft_predict):
                label = ft_label.split("~")[0]
                predict = ft_predict.split("~")[0]
                return label == predict


            def is_second_equal(ft_label, ft_predict):
                if "~" in ft_predict:
                    return ft_label == ft_predict
                else:
                    return ""


            sheet.cell(nrow, 5).value = is_first_equal(ft_label, llm_answer)
            sheet.cell(nrow, 6).value = is_second_equal(ft_label, llm_answer)

            sheet.cell(nrow, 7).value = sample['mannual_reason']
            sheet.cell(nrow, 8).value = sample['是否转接']
            sheet.cell(nrow, 9).value = sample['转接后技能组']
            online_first_ft = sample['AI一级FT']
            online_sec_ft = sample['AI二级FT'] if not isinstance(sample['AI二级FT'], float) else ""
            sheet.cell(nrow, 10).value = online_first_ft
            sheet.cell(nrow, 11).value = online_sec_ft

            sheet.cell(nrow, 12).value = ft_label.split("~")[0] == online_first_ft
            sheet.cell(nrow, 13).value = online_sec_ft == ft_label if not isinstance(sample['AI二级FT'], float) else ""
            sheet.cell(nrow, 14).value = sample['route_source']
            sheet.cell(nrow, 15).value = sample['是否标注']
            sheet.cell(nrow, 16).value = sample['标注前结果']

            nrow += 1
            if nrow % 2 == 0 and nrow != 0:
                workbook.save(OUTPUT_PATH.format(nrow))
                print(OUTPUT_PATH.format(nrow))
        except Exception as e:
            traceback.print_exc()
            print("Exception: ", e)
    workbook.save(OUTPUT_PATH.format(nrow))
