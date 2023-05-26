# -- coding: utf-8 --
# @Author : wuzixun
# @Time : 2023/5/19 11:26 AM
# @File : llm_ft_evaluete_e2e.py

from time import time
import torch
from transformers import AutoTokenizer, AutoModel
from peft import PeftModel
from openpyxl import Workbook
import json
from tqdm import tqdm
import pandas as pd
import traceback
from q_matching import Retrieval

#PEFT_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/output-question-20230520/"
PEFT_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/aidp-dialogue-intent-v2"
#CHATGLM_PATH = "/home/xiezizhe/wuzixun/LLM/chatglm-6b"
CHATGLM_PATH = "/home/xiezizhe/hanzhou/chatglm-6b"
#DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_ai_ft_eval/ai_ft_0524_一天全量_format.json"
DEV_DATA_PATH = "/home/xiezizhe/wuzixun/LLM/ChatGLM-Tuning/data/ks_ai_ft_eval/ai_ft_0524_一天全量_5原因_format.json"


def chatglm_inference(model, tokenizer, sample):
    context = f"Instruction: {sample['instruction']}\n"
    context += f"Input: {sample['input']}\n"
    context += "Answer: "

    response, _ = model.chat(tokenizer, context, history=[])
    return response


if __name__ == "__main__":

    tokenizer = AutoTokenizer.from_pretrained(CHATGLM_PATH, trust_remote_code=True)
    model = AutoModel.from_pretrained(CHATGLM_PATH, load_in_8bit=True, trust_remote_code=True, device_map='auto')
    model = model.eval()

    model = PeftModel.from_pretrained(model, PEFT_PATH)

    with open(DEV_DATA_PATH, "r", encoding='utf-8') as f:
        dev_data = json.load(f)

    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(1, 1).value = '对话'
    sheet.cell(1, 2).value = '首次关联工单功能树'
    sheet.cell(1, 3).value = 'LLM结果(微调后)'
    sheet.cell(1, 4).value = "LLM结果匹配人工Q"
    sheet.cell(1, 5).value = "LLM结果对应FT"
    sheet.cell(1, 6).value = "首次关联工单功能树和预测FT一致(一级FT)"
    sheet.cell(1, 7).value = "首次关联工单功能树和预测FT一致(二级FT)"
    sheet.cell(1, 8).value = "mannual_reason"
    sheet.cell(1, 9).value = "是否转接"
    sheet.cell(1, 10).value = "转接后技能组"
    sheet.cell(1, 11).value = "线上预测一级FT"
    sheet.cell(1, 12).value = "线上预测二级FT"
    sheet.cell(1, 13).value = "AI一级FT是否正确"
    sheet.cell(1, 14).value = "AI二级FT是否正确"

    with open("data/manual_ft_q_map/manual_q2ft.json", "r") as f:
        faq2ft = json.load(f)

    ret = Retrieval(load_from_disk=True, persist_directory=".chroma/biaozhunwen")

    nrow = 2
    for sample in tqdm(dev_data):
        try:
            sheet.cell(nrow, 1).value = sample['input']
            llm_answer = chatglm_inference(model, tokenizer, sample)
            ft_label = sample['output'].replace("~", "-")
            sheet.cell(nrow, 2).value = ft_label
            sheet.cell(nrow, 3).value = llm_answer

            retrieved_q = ret.retrieve(llm_answer)
            sheet.cell(nrow, 4).value = retrieved_q

            predict_ft = faq2ft[retrieved_q]
            sheet.cell(nrow, 5).value = predict_ft


            def is_first_equal(ft_label, ft_predict):
                label = ft_label.split("-")[0]
                predict = ft_predict.split("-")[0]
                return label == predict


            def is_second_equal(ft_label, ft_predict):
                return ft_label == ft_predict


            sheet.cell(nrow, 6).value = is_first_equal(ft_label, predict_ft)
            sheet.cell(nrow, 7).value = is_second_equal(ft_label, predict_ft)

            sheet.cell(nrow, 8).value = sample['mannual_reason']
            sheet.cell(nrow, 9).value = sample['是否转接']
            sheet.cell(nrow, 10).value = sample['转接后技能组']
            online_first_ft = sample['AI一级FT']
            online_sec_ft = sample['AI二级FT'].replace("~", "-") if not isinstance(sample['AI二级FT'], float) else ""
            sheet.cell(nrow, 11).value = online_first_ft
            sheet.cell(nrow, 12).value = online_sec_ft

            sheet.cell(nrow, 13).value = ft_label.split("-")[0] == online_first_ft
            sheet.cell(nrow, 14).value = online_sec_ft == ft_label if not isinstance(sample['AI二级FT'], float) else ""

            nrow += 1
            if nrow % 200 == 0 and nrow != 0:
                workbook.save(f"诉求LLM验证{nrow}_匹配人工Q_未筛选.xlsx")
                print(f"诉求LLM验证{nrow}_匹配人工Q_未筛选.xlsx: Save!")
        except Exception as e:
            traceback.print_exc()
            print("Exception: ", e)
    workbook.save(f"诉求LLM验证{nrow}_匹配人工Q_未筛选.xlsx")
